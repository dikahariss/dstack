#!/usr/bin/env python3
"""
test_pipeline.py — executable regressions for the defects that actually shipped.

    python test_pipeline.py            # needs ffmpeg; builds its own 6 s clip

Every test here corresponds to a defect that reached a delivered artifact. The
run_id test exists because 121 lines shipped with zero tests and the headline
fix did not work on the one table that motivated it — nobody wrote the four
lines that would have caught it.
"""
import json
import os
import shutil
import subprocess
import sys
import tempfile

import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
PY = sys.executable
FAILED = []


def run(*args, **kw):
    return subprocess.run([PY] + list(args), capture_output=True, text=True, **kw)


def check(name, cond, detail=""):
    print(f"  {'PASS' if cond else 'FAIL'}  {name}")
    if not cond:
        FAILED.append(f"{name}: {detail}")
        if detail:
            print(f"        {detail}")


def make_clip(path, seconds=6, audio=True):
    cmd = ["ffmpeg", "-v", "error", "-f", "lavfi",
           "-i", f"testsrc2=size=270x480:rate=25:duration={seconds}"]
    if audio:
        cmd += ["-f", "lavfi", "-i", f"sine=frequency=440:duration={seconds}"]
    cmd += ["-shortest", "-pix_fmt", "yuv420p", "-c:v", "libx264",
            "-preset", "ultrafast", path, "-y"]
    subprocess.run(cmd, check=True, capture_output=True)


def analyst_files(d, run_id, objective="awareness", score_mon=False):
    vid = pd.read_csv(f"{d}/video_master.csv").video_id.iloc[0]
    dur = pd.read_csv(f"{d}/video_master.csv").duration_s.iloc[0]
    base = dict(video_id=vid, run_id=run_id, coder_id="test")
    pd.DataFrame([{**base, "taxonomy_version": "3.1", "scored_at_utc": "2026-01-01T00:00:00Z",
                   "objective": objective, "funnel_stage": "top",
                   "subject_domain": "other", "intent": "entertain",
                   "production_mode": "single_take", "commercial_relationship": "none",
                   "platform_targets": "instagram_reels",
                   "pull_mechanisms": "novelty", "hook_device": "no_hook"}]
                 ).to_csv(f"{d}/semantic.csv", index=False)
    pd.DataFrame([{**base, "seg_idx": 1, "start_s": 0.0, "end_s": dur,
                   "segment_type": "hook_reveal", "visual_description": "x",
                   "retention_function": "y", "drop_risk": "low"}]
                 ).to_csv(f"{d}/segments.csv", index=False)
    rows = []
    for i in range(36):
        iid = f"XXX-{i:02d}"
        mon = iid in ("MON-01",)
        rows.append({**base, "item_id": iid, "persona": "P", "pillar": "Q",
                     "item": f"real text {i}", "weight": 2,
                     "applicable": "true", "score": 3 if i == 0 else 5,
                     "evidence": "e", "evidence_source": "computed", "action": ""})
    pd.DataFrame(rows).to_csv(f"{d}/scores.csv", index=False)


def main():
    if shutil.which("ffmpeg") is None:
        sys.exit("ffmpeg required")
    tmp = tempfile.mkdtemp(prefix="audit_test_")
    try:
        clip = os.path.join(tmp, "clip.mp4")
        make_clip(clip)

        print("\n[1] extractor")
        a = os.path.join(tmp, "a")
        r = run(f"{HERE}/extract_features.py", clip, a)
        check("runs and writes a master row", os.path.exists(f"{a}/video_master.csv"),
              r.stderr[-300:])
        m = pd.read_csv(f"{a}/video_master.csv").iloc[0]

        # no sidecar -> limitation must fire, join columns must be NULL not ""
        lim = open(f"{a}/limitations.txt").read()
        check("no sidecar -> limitation emitted", "sidecar" in lim.lower(), lim[:200])
        check("no sidecar -> platform_post_id is NULL",
              pd.isna(m.platform_post_id) or str(m.platform_post_id).strip() == "")

        # foreign JSON sibling must NOT be harvested
        b = os.path.join(tmp, "b")
        clip2 = os.path.join(tmp, "clip2.mp4")
        shutil.copy(clip, clip2)
        json.dump({"note": "my own notes"}, open(os.path.join(tmp, "clip2.json"), "w"))
        run(f"{HERE}/extract_features.py", clip2, b)
        lim2 = open(f"{b}/limitations.txt").read()
        m2 = pd.read_csv(f"{b}/video_master.csv").iloc[0]
        check("foreign .json is NOT consumed", str(m2.sidecar_file or "") in ("", "nan"),
              f"sidecar_file={m2.sidecar_file!r}")
        check("foreign .json still emits the missing-join limitation",
              "sidecar" in lim2.lower())

        # real yt-dlp sidecar IS harvested
        c = os.path.join(tmp, "c")
        clip3 = os.path.join(tmp, "clip3.mp4")
        shutil.copy(clip, clip3)
        json.dump({"id": "ABC123", "extractor_key": "Instagram",
                   "webpage_url": "https://example/p/ABC123/", "uploader_id": "999",
                   "channel": "acct", "timestamp": 1700000000, "epoch": 1700000900,
                   "like_count": 10, "comment_count": 2,
                   "description": "hello #one #two"},
                  open(os.path.join(tmp, "clip3.info.json"), "w"))
        run(f"{HERE}/extract_features.py", clip3, c)
        m3 = pd.read_csv(f"{c}/video_master.csv").iloc[0]
        check("yt-dlp sidecar harvested", m3.platform_post_id == "ABC123")
        check("capture time uses epoch, not mtime",
              str(m3.engagement_captured_at_utc).startswith("2023-11-14"),
              str(m3.engagement_captured_at_utc))
        check("caption + hashtags harvested",
              "hello" in str(m3.caption) and "one" in str(m3.hashtags))
        check("schema_version bumped with the new columns", str(m3.schema_version) == "3.1",
              str(m3.schema_version))

        print("\n[2] validator")
        analyst_files(a, "run-a")
        r = run(f"{HERE}/validate_audit.py", a)
        check("empty action on a non-5 item is caught",
              "ACTION" in r.stdout or r.returncode == 1, r.stdout[:200])
        sc = pd.read_csv(f"{a}/scores.csv")
        sc["action"] = "do x at 0:01"
        sc.to_csv(f"{a}/scores.csv", index=False)
        r = run(f"{HERE}/validate_audit.py", a)
        check("clean audit validates", r.returncode == 0, r.stdout[-400:])
        # illegal enum
        sem = pd.read_csv(f"{a}/semantic.csv"); sem.loc[0, "objective"] = "sell"
        sem.to_csv(f"{a}/semantic.csv", index=False)
        r = run(f"{HERE}/validate_audit.py", a)
        check("illegal enum value fails", r.returncode == 1 and "ENUM" in r.stdout,
              r.stdout[:200])
        sem.loc[0, "objective"] = "awareness"; sem.to_csv(f"{a}/semantic.csv", index=False)

        print("\n[3] corpus keeps two codings")
        d2 = os.path.join(tmp, "a2")
        shutil.copytree(a, d2)
        for t in ("semantic.csv", "segments.csv", "scores.csv"):
            df = pd.read_csv(f"{d2}/{t}"); df["run_id"] = "run-b"
            df.to_csv(f"{d2}/{t}", index=False)
        corp = os.path.join(tmp, "corpus")
        run(f"{HERE}/merge_corpus.py", corp, a)
        run(f"{HERE}/merge_corpus.py", corp, d2)
        for tbl, n in [("corpus_scores.csv", 72), ("corpus_segments.csv", 2),
                       ("corpus_semantic.csv", 2)]:
            p = os.path.join(corp, tbl)
            got = len(pd.read_csv(p)) if os.path.exists(p) else 0
            check(f"{tbl} keeps both codings ({n} rows)", got == n, f"got {got}")
        sem_c = pd.read_csv(os.path.join(corp, "corpus_semantic.csv"))
        check("corpus_semantic carries both run_ids",
              set(sem_c.run_id) == {"run-a", "run-b"}, str(set(sem_c.run_id)))

        print("\n[4] missing run_id warns loudly")
        d3 = os.path.join(tmp, "a3")
        shutil.copytree(a, d3)
        for t in ("semantic.csv", "segments.csv", "scores.csv"):
            df = pd.read_csv(f"{d3}/{t}").drop(columns=["run_id"])
            df.to_csv(f"{d3}/{t}", index=False)
        r = run(f"{HERE}/merge_corpus.py", os.path.join(tmp, "corpus2"), d3, "--strict")
        check("no run_id -> warning printed", "run_id" in r.stdout, r.stdout[:200])
        check("no run_id -> --strict exits non-zero", r.returncode != 0)

        print("\n[5] --forget removes every row")
        vid = pd.read_csv(f"{a}/video_master.csv").video_id.iloc[0]
        r = run(f"{HERE}/merge_corpus.py", corp, "--forget", vid)
        left = sum(len(pd.read_csv(os.path.join(corp, f)))
                   for f in os.listdir(corp) if f.startswith("corpus_") and f.endswith(".csv"))
        check("--forget empties the corpus", left == 0, f"{left} rows left")

        print("\n[6] workbook column widths follow names, not positions")
        analyst_files(a, "run-a")
        sc = pd.read_csv(f"{a}/scores.csv"); sc["action"] = "do x"; sc.to_csv(f"{a}/scores.csv", index=False)
        pd.DataFrame([{"video_id": vid, "run_id": "run-a", "coder_id": "t",
                       "evidence_source": "vision", "t_start_s": 0, "t_end_s": 1,
                       "text": "hello", "text_role": "caption",
                       "position_band": "center", "language": "en",
                       "is_burned_in": "true"}]).to_csv(f"{a}/onscreen_text.csv", index=False)
        r = run(f"{HERE}/build_workbook.py", a)
        ok = os.path.exists(f"{a}/video_audit.xlsx")
        check("workbook builds", ok, r.stdout[-300:] + r.stderr[-300:])
        if ok:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter as L
            ws = load_workbook(f"{a}/video_audit.xlsx")["OnScreen_Text"]
            hdr = [c.value for c in ws[1]]
            w = ws.column_dimensions[L(hdr.index("text") + 1)].width
            check("free-text column is wide, not 12", w and w >= 40, f"width={w}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    print(f"\n{'ALL PASS' if not FAILED else str(len(FAILED)) + ' FAILED'}")
    for f in FAILED:
        print(f"  - {f}")
    return 1 if FAILED else 0


if __name__ == "__main__":
    sys.exit(main())
