#!/usr/bin/env python3
"""
build_workbook.py (v3) — assemble the audit Excel workbook.

Usage:
    python build_workbook.py <audit_dir> [--out <workbook.xlsx>]

<audit_dir> is the output of extract_features.py. Analyst-written files are read
from that SAME directory — v2 took --scores/--segments pointing at an undefined
<workdir>, which contradicted the two steps that wrote them into <audit_dir>.

Files read (all optional except the extractor output):
    scores.csv           video_id, item_id, persona, pillar, item, weight,
                         applicable, score, evidence, action
    segments.csv         video_id, seg_idx, start_s, end_s, segment_type,
                         visual_description, retention_function, drop_risk
    semantic.csv         one row, see references/schema.md
    recommendations.csv  video_id, rec_idx, item_ids, recommendation, effort,
                         expected_metric, expected_direction, personas_lifted

These column lists are asserted at load, and they are the SAME lists that
references/schema.md declares. v2's docstring named a 7-column scores.csv and a
6-column segments.csv while its code required 9 and 8.

HEALTH INDEX — the definition, stated here because the report has to quote it:
    weighted        = weight * score                (per applicable item)
    weighted_max    = weight * 5                    (per applicable item)
    health_index    = 100 * SUM(weighted) / SUM(weighted_max)
Items with applicable=false are excluded from BOTH sums. v2 had no applicable
flag, so an item omitted as irrelevant silently shrank the denominator and
raised the index — the same video could read 58.7 or 63.7.

Only LibreOffice-safe formulas are used (SUM/SUMIF/COUNTIF/IF/IFERROR/ROUND).
"""
import argparse
import json
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

HEAD = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY = Font(name="Arial", size=10)
FILL = PatternFill("solid", fgColor="1F3864")

SCORES_COLUMNS = ["video_id", "run_id", "coder_id", "item_id", "persona", "pillar", "item", "weight",
                  "applicable", "score", "evidence", "evidence_source", "action"]
SEGMENTS_COLUMNS = ["video_id", "run_id", "coder_id", "seg_idx", "start_s", "end_s", "segment_type",
                    "visual_description", "retention_function", "drop_risk"]
TEXT_COLUMNS = ["video_id", "run_id", "coder_id", "evidence_source", "t_start_s", "t_end_s",
                "text", "text_role", "position_band", "language", "is_burned_in"]
RECS_COLUMNS = ["video_id", "run_id", "coder_id", "rec_idx", "item_ids", "recommendation", "effort",
                "expected_metric", "expected_direction", "personas_lifted"]


def _w(columns, **overrides):
    """Widths keyed by COLUMN NAME, not position. Positional lists silently
    shifted every width by two when run_id/coder_id were inserted, collapsing
    `text`, `visual_description` and `recommendation` to 11-20 chars while
    numeric columns got 48-52."""
    default = {"video_id": 15, "run_id": 18, "coder_id": 15, "item_id": 9}
    return [overrides.get(c, default.get(c, 12)) for c in columns]


def style_sheet(ws, ncols, widths=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font, cell.fill = HEAD, FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        w = (widths[c - 1] if widths and c - 1 < len(widths) else None)
        ws.column_dimensions[get_column_letter(c)].width = (
            w if w else max(12, min(46, len(str(cell.value or "")) + 4)))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def write_df(wb, title, df, widths=None):
    ws = wb.create_sheet(title)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    style_sheet(ws, df.shape[1], widths)
    return ws


def load_checked(path, required, label, optional=()):
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        return None
    df = pd.read_csv(path, dtype={"video_id": "string", "item_id": "string"})
    required = [c for c in required if c not in optional or c in df.columns]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise SystemExit(f"{label} ({path}) missing columns: {missing}\n"
                         f"Required, per references/schema.md: {required}")
    return df[required]


def fmt(v):
    """Render a possibly-missing measurement. NULL is not zero, and the workbook
    must not let a reader mistake one for the other."""
    return "not measured" if v is None or (isinstance(v, float) and pd.isna(v)) else v


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("audit_dir")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    ad = args.audit_dir
    out = args.out or os.path.join(ad, "video_audit.xlsx")

    s = json.load(open(f"{ad}/analysis.json"))
    s = s.get("master", s)
    tl = pd.read_csv(f"{ad}/timeline_per_second.csv")
    sh = pd.read_csv(f"{ad}/shot_list.csv")
    scores = load_checked(f"{ad}/scores.csv", SCORES_COLUMNS, "scores.csv", ("run_id","coder_id"))
    if scores is None:
        raise SystemExit(f"{ad}/scores.csv not found. Score the checklist first "
                         "(references/persona_checklist.md), then re-run.")
    segs = load_checked(f"{ad}/segments.csv", SEGMENTS_COLUMNS, "segments.csv", ("run_id","coder_id"))
    recs = load_checked(f"{ad}/recommendations.csv", RECS_COLUMNS, "recommendations.csv", ("run_id","coder_id"))
    otext = load_checked(f"{ad}/onscreen_text.csv", TEXT_COLUMNS, "onscreen_text.csv", ("run_id","coder_id"))

    scores["applicable"] = (scores["applicable"].astype(str).str.strip().str.lower()
                            .isin(("true", "1", "yes", "y")))

    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("PROVENANCE", ""),
        ("video_id", s["video_id"]), ("File", s["source_file"]),
        ("Extracted (UTC)", s.get("extracted_at_utc")),
        ("Schema / extractor", f"{s.get('schema_version')} / {s.get('extractor_version')}"),
        ("Target platform", s.get("param_platform")),
        ("Sampling rate (fps)", s.get("param_sample_fps")),
        ("Scene threshold", s.get("param_scene_threshold")),
        ("OCR available", s.get("ocr_available")),
        ("Face detection available", s.get("face_detection_available")),
        ("Motion comparable across videos", s.get("motion_comparable")),
        ("", ""),
        ("TECHNICAL", ""),
        ("Duration (s)", s["duration_s"]),
        ("Resolution", f"{s['width']}x{s['height']}"),
        ("Aspect ratio", s["aspect_ratio"]), ("Orientation", s.get("orientation")),
        ("FPS", fmt(s.get("fps"))), ("Video codec", s["video_codec"]),
        ("Audio codec", fmt(s.get("audio_codec"))),
        ("Bitrate (kbps)", fmt(s.get("bitrate_kbps"))),
        ("File size (MB)", fmt(s.get("filesize_mb"))),
        ("", ""),
        ("EDITING", ""),
        ("Shots", s.get("n_shots")), ("Total cuts", s["total_cuts"]),
        ("Cuts per second", s["cuts_per_second"]),
        ("Shot duration median (s)", s.get("shot_median_s")),
        ("Shots under 0.5 s", f"{s.get('shot_under_0_5s_n')} of {s.get('n_shots')} "
                              f"({s.get('shot_pct_under_0_5s')}%)"),
        ("Cut-on-beat", f"{fmt(s.get('n_cuts_synced'))} of {fmt(s.get('n_cuts_tested'))} "
                        f"({fmt(s.get('cut_beat_sync_pct'))}%)"),
        ("  random baseline (%)", fmt(s.get("cut_beat_sync_baseline_pct"))),
        ("  lift (pp)", fmt(s.get("cut_beat_sync_lift"))),
        ("  p-value (lift vs chance)", fmt(s.get("cut_beat_sync_p"))),
        ("", ""),
        ("VISUAL", ""),
        ("Mean brightness (0-1)", s["brightness_mean"]),
        ("Mean saturation (0-1)", s["saturation_mean"]),
        ("Colorfulness mean / median", f"{s['colorfulness_mean']} / {s.get('colorfulness_median')}"),
        ("Sharpness mean / median", f"{s['sharpness_mean']} / {s.get('sharpness_median')}"),
        ("Motion mean / median", f"{fmt(s.get('motion_mean'))} / {fmt(s.get('motion_median'))}"),
        ("Frames with face (%)", fmt(s.get("frames_with_face_pct"))),
        ("On-screen text (% of OCR frames)", fmt(s.get("seconds_with_text_pct"))),
        ("  OCR frames with text / sampled",
         f"{fmt(s.get('ocr_frames_with_text'))} / {fmt(s.get('ocr_frames_sampled'))}"),
        ("Loop similarity (0-1)", s["loop_similarity"]),
        ("Edge energy in safe zone (share)", s.get("edge_energy_safe")),
        ("  safe-zone area (share of frame)", s.get("safe_area_fraction")),
        ("  index (1.0 = uniform texture)", s.get("edge_energy_safe_index")),
        ("", ""),
        ("AUDIO", ""),
        ("Integrated loudness (LUFS)", fmt(s.get("integrated_lufs"))),
        ("Loudness range (LU)", fmt(s.get("loudness_range_lu"))),
        ("True peak (dBFS)", fmt(s.get("true_peak_dbfs"))),
        ("Estimated BPM", fmt(s.get("bpm_estimate"))),
        ("Silence segments", fmt(s.get("n_silences"))),
    ]
    ws.append(["Metric", "Value"])
    for k, v in rows:
        ws.append([k, v])
    style_sheet(ws, 2, [40, 30])
    for row in ws.iter_rows(min_row=2):
        if row[1].value == "" and row[0].value:
            row[0].font = Font(name="Arial", size=10, bold=True)
    n = ws.max_row + 2
    ws.cell(n, 1, "Numbers above are computed from the file (ffmpeg/OpenCV"
                  + (", Tesseract" if s.get("ocr_available") else "")
                  + "). 'not measured' means the measurement could not run on "
                    "this machine — it does NOT mean zero.").font = \
        Font(name="Arial", size=9, italic=True)

    # ---- Audit checklist ----
    ws = wb.create_sheet("Audit_Checklist")
    ws.append(["video_id", "item_id", "Persona", "Pillar", "Item", "Weight (1-3)",
               "Applicable", "Score (0-5)", "Evidence", "Evidence source", "Action",
               "Weighted", "Max", "Status"])
    for i, r in enumerate(scores.itertuples(index=False), start=2):
        ws.append(list(r))
        # G = Applicable, F = Weight, H = Score
        ws.cell(i, 12, f'=IF(G{i}=TRUE,F{i}*H{i},"")')
        ws.cell(i, 13, f'=IF(G{i}=TRUE,F{i}*5,"")')
        ws.cell(i, 14, f'=IF(G{i}<>TRUE,"N/A",IF(H{i}<=1,"CRITICAL",'
                       f'IF(H{i}<=3,"NEEDS WORK","GOOD")))')
    n = len(scores) + 1
    ws.cell(n + 2, 1, "CONTENT HEALTH INDEX").font = Font(name="Arial", bold=True, size=11)
    ws.cell(n + 3, 1, "Applicable items"); ws.cell(n + 3, 12, f'=COUNTIF(G2:G{n},TRUE)')
    ws.cell(n + 4, 1, "Excluded (N/A)"); ws.cell(n + 4, 12, f'=COUNTIF(G2:G{n},FALSE)')
    ws.cell(n + 5, 1, "Weighted total"); ws.cell(n + 5, 12, f"=SUM(L2:L{n})")
    ws.cell(n + 6, 1, "Weighted max"); ws.cell(n + 6, 12, f"=SUM(M2:M{n})")
    ws.cell(n + 7, 1, "Health index (%)")
    ws.cell(n + 7, 12, f"=IFERROR(ROUND(L{n+5}/L{n+6}*100,1),\"\")")
    ws.cell(n + 8, 1, "CRITICAL items"); ws.cell(n + 8, 12, f'=COUNTIF(N2:N{n},"CRITICAL")')
    ws.cell(n + 9, 1, "GOOD items"); ws.cell(n + 9, 12, f'=COUNTIF(N2:N{n},"GOOD")')
    ws.cell(n + 11, 1, "Index = 100 * SUM(weight*score) / SUM(weight*5), over "
                       "APPLICABLE items only. It is a weighted mean of ordinal "
                       "analyst judgments, not a measurement. Do not compare it "
                       "across videos with different applicable sets."
            ).font = Font(name="Arial", size=9, italic=True)
    style_sheet(ws, 14, [15, 9, 20, 14, 34, 9, 10, 9, 42, 13, 40, 10, 8, 13])
    for row in ws.iter_rows(min_row=n + 2, max_row=n + 11):
        for cell in row:
            cell.font = BODY
    ws.conditional_formatting.add(f"H2:H{n}", CellIsRule(
        operator="lessThanOrEqual", formula=["1"],
        fill=PatternFill("solid", fgColor="F8CBAD")))
    ws.conditional_formatting.add(f"H2:H{n}", CellIsRule(
        operator="greaterThanOrEqual", formula=["4"],
        fill=PatternFill("solid", fgColor="C6E0B4")))

    # ---- Per-persona rollup ----
    ws2 = wb.create_sheet("Score_per_Persona")
    ws2.append(["Persona", "Items", "Applicable", "Weighted", "Max", "Index (%)"])
    for p in scores.persona.unique():
        r = ws2.max_row + 1
        sub = scores[scores.persona == p]
        ws2.append([p, int(len(sub)), int(sub.applicable.sum()), None, None, None])
        ws2.cell(r, 4, f"=SUMIF(Audit_Checklist!C2:C{n},A{r},Audit_Checklist!L2:L{n})")
        ws2.cell(r, 5, f"=SUMIF(Audit_Checklist!C2:C{n},A{r},Audit_Checklist!M2:M{n})")
        ws2.cell(r, 6, f'=IFERROR(ROUND(D{r}/E{r}*100,1),"")')
    style_sheet(ws2, 6, [28, 8, 11, 12, 10, 12])

    # ---- Optional analyst sheets ----
    sem_p = os.path.join(ad, "semantic.csv")
    if os.path.exists(sem_p) and os.path.getsize(sem_p) > 0:
        sem = pd.read_csv(sem_p).T.reset_index()
        sem.columns = ["Field", "Value"]
        write_df(wb, "Semantic", sem, widths=[26, 70])
    if recs is not None:
        write_df(wb, "Recommendations", recs, widths=_w(RECS_COLUMNS, recommendation=60, item_ids=16, personas_lifted=26))
    if segs is not None:
        write_df(wb, "Narrative_Segments", segs, widths=_w(SEGMENTS_COLUMNS, visual_description=52, retention_function=40, segment_type=18))

    if otext is not None:
        write_df(wb, "OnScreen_Text", otext, widths=_w(TEXT_COLUMNS, text=60, evidence_source=14))

    write_df(wb, "Timeline_per_second", tl)
    write_df(wb, "Shot_list", sh)
    fr_p = os.path.join(ad, "frame_features.csv")
    if os.path.exists(fr_p):
        write_df(wb, "Frame_features", pd.read_csv(fr_p, dtype={"video_id": "string"}))

    lim = ""
    lp = os.path.join(ad, "limitations.txt")
    if os.path.exists(lp):
        lim = open(lp).read()
    wsl = wb.create_sheet("Limitations")
    wsl.append(["What this audit could not measure"])
    for line in lim.splitlines():
        wsl.append([line])
    style_sheet(wsl, 1, [120])

    dd = pd.DataFrame([
        ("brightness", "0-1", "Mean luminance; <0.35 dark, >0.65 bright"),
        ("contrast", "0-1", "Luminance std dev; visual 'punch' proxy"),
        ("saturation", "0-1", "Mean HSV saturation; low = muted palette"),
        ("colorfulness", "0-100+", "Hasler-Susstrunk index. Validated against "
                                   "perceived colorfulness of natural images — NOT "
                                   "against scroll-stopping."),
        ("sharpness", "variance", "Laplacian variance; heavy right tail, read the median"),
        ("motion_score", "0-255", "Mean abs difference to the PREVIOUS sampled frame. "
                                  "Only comparable at equal param_motion_dt_s. Null on frame 0."),
        ("edge_density", "0-1", "Canny edge-pixel share; visual busyness proxy"),
        ("face_count", "count", "Frontal Haar cascade. Null when unavailable, never 0."),
        ("dominant_color", "hex", "k-means (k=3) top cluster, seeded RNG"),
        ("cuts_in_sec", "count", "Scene cuts landing in that second, at param_scene_threshold"),
        ("audio_dbfs", "dBFS", "RMS level per 0.5 s window"),
        ("speech_band_ratio", "0-1", "Energy share 300-3400 Hz. Music occupies this "
                                     "band too — this does NOT detect speech."),
        ("bass_ratio", "0-1", "Energy share <250 Hz"),
        ("ocr_text", "-", "Tesseract, conf>55, ~1 fps. Null when OCR unavailable."),
        ("loop_similarity", "0-1", "HSV-hist correlation, first vs last SAMPLED frame. "
                                   "Composition-blind: same palette scores high."),
        ("cut_beat_sync_lift", "pp", "sync% minus the union coverage of beat windows. "
                                     "Read cut_beat_sync_p before calling it craft."),
        ("edge_energy_safe_index", "ratio", "Edge energy in the safe box / safe-box area. "
                                            "1.0 = what uniform texture alone would give."),
    ], columns=["Column", "Unit", "Definition"])
    write_df(wb, "Data_Dictionary", dd, widths=[26, 12, 70])

    wb.save(out)
    print(f"Saved {out}\nSheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
